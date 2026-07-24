"""DOL/OFLC LCA historical sponsorship enrichment.

A second, distinct signal from `app.visa_scan` -- that module only ever reads
one specific job posting's own text ("does this posting say they'll
sponsor"); this module reads the Department of Labor's public LCA (H-1B,
H-1B1, E-3) disclosure data to answer a different question entirely: "has
this company actually filed for sponsorship before, and when."

DOL/OFLC releases these files quarterly (cumulative within the federal fiscal
year, Oct 1 - Sep 30) as .xlsx, ~100-600MB depending on the quarter, at
https://www.dol.gov/agencies/eta/foreign-labor/performance -- there is no API,
and the site's bot protection blocks unattended automated downloads (confirmed
live: both a plain HTTP request and this project's own sandboxed browser tool
were blocked). Downloading the quarterly file is therefore a manual, one-time
step per quarter -- run `python -m app.cli lca-enrich <path-to-xlsx>` after
downloading it yourself.

Real, confirmed data quality note: DOL's own site claims EMPLOYER_FEIN is
redacted as PII, but the live FY2026 Q2 file actually has it populated for
every row checked. Not used here regardless -- this project's `companies`
table has no FEIN column, and name-matching (the same word-boundary,
case/whitespace-insensitive `_normalize()` already used everywhere else in
this codebase) is the only matching signal available today. First real test
against this project's live 2,682-company table found 223 clean matches (no
false positives spotted, including in the highest-collision-risk short-name
group) -- see RUNBOOK.md for the full investigation.
"""
from __future__ import annotations

import sqlite3
from datetime import datetime

from app.companies import _normalize

CASE_STATUS_COL = "CASE_STATUS"
DECISION_DATE_COL = "DECISION_DATE"
EMPLOYER_NAME_COL = "EMPLOYER_NAME"


def _extract_certifications(rows) -> dict[str, tuple[datetime | None, str, str]]:
    """rows: iterable of (case_status, decision_date, employer_name) tuples,
    as they appear in a DOL LCA disclosure file (any order/status mix, not
    pre-filtered). Returns normalized_employer_name -> (decision_date,
    raw_employer_name, status) for the single most recent "Certified"/
    "Certified - Withdrawn" record per employer -- multiple LCAs per employer
    per file is the norm, only the latest matters for "when did they last
    sponsor."""
    best: dict[str, tuple[datetime | None, str, str]] = {}
    for status, decision_date, raw_name in rows:
        if not raw_name or not status or "Certified" not in status:
            continue
        norm = _normalize(raw_name)
        existing = best.get(norm)
        if existing is None or (decision_date and (not existing[0] or decision_date > existing[0])):
            best[norm] = (decision_date, raw_name, status)
    return best


def merge_lca_data(
    *parsed: dict[str, tuple[datetime | None, str, str]],
) -> dict[str, tuple[datetime | None, str, str]]:
    """Combines the results of multiple `parse_lca_disclosure_file` calls
    (e.g. this fiscal year's file plus one or more prior fiscal years'
    final/annual files) into one lookup, keeping the most recent decision
    date per employer across all of them.

    Each DOL quarterly file is cumulative *within its own fiscal year* --
    FY2026 Q2 already contains everything FY2026 Q1 does, so passing every
    quarter of the same year is redundant, not additive. The genuinely new
    coverage comes from adding a *different fiscal year's* file (typically
    that year's Q4/final release, its most complete one)."""
    merged: dict[str, tuple[datetime | None, str, str]] = {}
    for data in parsed:
        for norm, (decision_date, raw_name, status) in data.items():
            existing = merged.get(norm)
            if existing is None or (decision_date and (not existing[0] or decision_date > existing[0])):
                merged[norm] = (decision_date, raw_name, status)
    return merged


def parse_lca_disclosure_file(path: str) -> dict[str, tuple[datetime | None, str, str]]:
    """Reads a real DOL OFLC LCA disclosure .xlsx file. Looks up columns by
    header name, not position -- DOL changed the record layout once already
    (FY2020), and probably will again."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    col_idx = {name: i for i, name in enumerate(header) if name}

    missing = [c for c in (CASE_STATUS_COL, DECISION_DATE_COL, EMPLOYER_NAME_COL) if c not in col_idx]
    if missing:
        raise ValueError(f"LCA file missing expected column(s): {missing}")

    status_i = col_idx[CASE_STATUS_COL]
    decision_i = col_idx[DECISION_DATE_COL]
    employer_i = col_idx[EMPLOYER_NAME_COL]

    def _rows():
        for row in rows_iter:
            yield row[status_i], row[decision_i], row[employer_i]

    return _extract_certifications(_rows())


def run_lca_enrichment(conn: sqlite3.Connection, parsed: dict[str, tuple], main_ws=None) -> dict:
    """Matches every tracked company against already-parsed LCA data by
    normalized name and updates `dol_lca_employer_name`/
    `last_lca_certified_date` for each match. `parsed` is the dict returned by
    `parse_lca_disclosure_file`/`_extract_certifications` -- kept as a
    separate argument (not a file path) so this function stays trivially
    testable and so multiple quarterly files can be merged (via repeated
    `_extract_certifications` calls picking the latest date) before a single
    DB update pass.

    If `main_ws` is given, also pushes the match onto every one of that
    company's jobs currently on Beacon -- without this, a company matched
    after its job's Beacon row already exists would never show the new
    columns there (same gap app.enrichment._push_to_beacon exists to close
    for FMP/StartupHub, reused here directly).

    Never regresses a company's stored last_lca_certified_date -- if the
    date already in the DB is more recent than this hit's date (e.g. an
    older fiscal year's file was passed in a run that didn't also include
    whatever newer file produced the currently-stored date), the existing
    value is left untouched rather than being overwritten backward in time."""
    from app.sheets import update_company_columns

    companies = conn.execute("SELECT * FROM companies").fetchall()
    matched = 0
    for company in companies:
        hit = parsed.get(_normalize(company["name"]))
        if hit is None:
            continue
        decision_date, raw_name, _status = hit

        existing_date_str = company["last_lca_certified_date"]
        if existing_date_str is not None:
            existing_date = datetime.fromisoformat(existing_date_str)
            if decision_date is None or existing_date > decision_date:
                continue

        conn.execute(
            "UPDATE companies SET dol_lca_employer_name = ?, last_lca_certified_date = ? WHERE id = ?",
            (raw_name, decision_date.isoformat() if decision_date else None, company["id"]),
        )
        conn.commit()
        matched += 1

        if main_ws is not None:
            updated_company = conn.execute(
                "SELECT * FROM companies WHERE id = ?", (company["id"],)
            ).fetchone()
            job_ids = conn.execute(
                "SELECT id FROM jobs WHERE company_id = ? AND sheet_row_number IS NOT NULL",
                (company["id"],),
            ).fetchall()
            for job_row in job_ids:
                update_company_columns(main_ws, job_row["id"], updated_company)

    return {"companies_checked": len(companies), "matched": matched}
